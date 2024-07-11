import argparse
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from matplotlib.colors import ListedColormap


def sigmoid(x: int | float | np.integer | np.floating | list | np.ndarray):

    if isinstance(x, (int, float, np.integer, np.floating)):
        return np.exp(x) / (1 + np.exp(x))

    elif isinstance(x, list):
        result = np.array(x)
        shape = result.shape
        return np.fromiter(map(lambda inx: np.exp(inx) / (1 + np.exp(inx)), result.flatten()), dtype=np.float32).reshape(shape).tolist()

    elif isinstance(x, np.ndarray):
        shape = x.shape
        return np.fromiter(map(lambda inx: np.exp(inx) / (1 + np.exp(inx)), x.flatten()), dtype=np.float32).reshape(shape)

    else:
        raise TypeError("Please enter the correct format !!!")


def check(vertexes, neighbor_list):
    for idx, v in enumerate(vertexes):
        index = neighbor_list[idx]
        if v in vertexes[index]:
            return False
    return True


def generate_neighbor(size: tuple, graph_mode="King"):
    neighbor_list = []

    for i in range(size[0]):
        for j in range(size[1]):
            if i == 0 and j == 0:  # top left
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i + 1) + j + 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * i + j + 1])

            elif i == 0 and j == size[1] - 1:  # top right
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i + 1) + j - 1, size[1] * i + j - 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * i + j - 1])

            elif i == size[0] - 1 and j == 0:  # bottom left
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * (i - 1) + j + 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * i + j + 1])

            elif i == size[0] - 1 and j == size[1] - 1:  # bottom right
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * (i - 1) + j - 1, size[1] * i + j - 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * i + j - 1])

            elif i == 0 and (0 < j < size[1] - 1):  # top edge
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i + 1) + j - 1, size[1] * (i + 1) + j + 1, size[1] * i + j - 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * i + j - 1, size[1] * i + j + 1])

            elif i == size[0] - 1 and (0 < j < size[1] - 1):  # bottom edge
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * (i - 1) + j - 1, size[1] * (i - 1) + j + 1, size[1] * i + j - 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i - 1) + j, size[1] * i + j - 1, size[1] * i + j + 1])

            elif (0 < i < size[0] - 1) and j == 0:  # left edge
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i - 1) + j, size[1] * (i + 1) + j + 1, size[1] * (i - 1) + j + 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i - 1) + j, size[1] * i + j + 1])

            elif (0 < i < size[0] - 1) and j == size[1] - 1:  # right edge
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i - 1) + j, size[1] * (i + 1) + j - 1, size[1] * (i - 1) + j - 1, size[1] * i + j - 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i - 1) + j, size[1] * i + j - 1])

            else:  # inside
                if graph_mode in ["k", "K", "king", "King"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i + 1) + j - 1, size[1] * (i + 1) + j + 1, size[1] * (i - 1) + j, size[1] * (i - 1) + j - 1, size[1] * (i - 1) + j + 1, size[1] * i + j - 1, size[1] * i + j + 1])
                elif graph_mode in ["l", "L", "lattice", "Lattice"]:
                    neighbor_list.append([size[1] * (i + 1) + j, size[1] * (i - 1) + j, size[1] * i + j - 1, size[1] * i + j + 1])

    return neighbor_list


def generate_coupling(spin, neighbor, mode="Ising"):
    coupling = []
    for n in range(len(neighbor)):
        if mode in ["Ising", "ising", "I", "i"]:
            coupling.append((np.int32(spin[neighbor[n]] != spin[n]) * 2 - 1).tolist())

        elif mode in ["Potts", "potts", "P", "p"]:
            coupling.append([1 for _ in neighbor[n]])

        else:
            raise NameError('Input Ising or Potts for parameter "mode".')
    return coupling


def generate_temperature(start, stop, num):
    if num < max(start / stop, stop / start):
        raise ValueError("Set a larger parameter num!")
    else:
        rank = int(np.log2(max(start / stop, stop / start)) + 1)
        repeat_size = num // rank
        temperature = np.array([stop]).repeat(num)

        if stop >= start:
            temperature[:repeat_size * rank] = np.array([start * 2 ** r for r in range(rank)]).repeat(repeat_size)
        else:
            temperature[:repeat_size * rank] = np.array([start / 2 ** r for r in range(rank)]).repeat(repeat_size)
        return temperature


def load_custom(path, rows: (int, int), cols: (int, int)):
    spins = np.array([])
    start_row, end_row = rows
    start_col, end_col = cols
    sheet = load_workbook(path).active
    for i in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for j in i:
            spins = np.append(spins, j.value)

    return spins


class Potts2D:
    def __init__(self, spin_num, spin_value, couplings, neighbors: list, temperature, MC_steps, Ising_mode=False):
        self.n = spin_num
        self.q = spin_value
        self.cps = couplings
        self.nbs = neighbors
        self.tpt = temperature
        self.mc_steps = MC_steps
        self.Ising_mode = Ising_mode

        # spin initialization
        if self.Ising_mode:
            self.spin = np.random.choice([-1, 1], size=self.n)

        else:
            self.spin = np.zeros(shape=self.n, dtype=np.int32)
            # self.spin = np.random.choice(self.q, size=self.n)

        self.spin_log = self.spin
        self.energy_log = [Potts2D.sys_energy(self.spin, self.cps, self.nbs)]

    @staticmethod
    def loc_energy(spin, coupling, neighbor: list):
        return np.sum(np.array(coupling) * (2 * np.int32(np.array(neighbor) == spin) - 1))

    @staticmethod
    def sys_energy(spins, couplings, neighbors):
        energy = 0
        for n, (spin, neighbor) in enumerate(zip(spins, neighbors)):
            energy += Potts2D.loc_energy(spin, couplings[n], spins[neighbor])
        return energy

    def update(self, temperature):
        for n, (spin, coupling, index) in enumerate(zip(self.spin, self.cps, self.nbs)):
            neighbor = self.spin[index]
            spin_stochastic = np.random.choice(self.q[self.q != spin])
            delta_energy = Potts2D.loc_energy(spin_stochastic, coupling, neighbor) - Potts2D.loc_energy(spin, coupling, neighbor)

            # spin update
            # if delta_energy < 0 or sigmoid(delta_energy * temperature) < np.random.uniform():
            if delta_energy < 0 or np.exp(- delta_energy / temperature) > np.random.uniform():
                self.spin[n] = spin_stochastic

    def run(self):
        # for t in np.linspace(start=self.tpt[0], stop=self.tpt[1], num=self.mc_steps):
        for t in generate_temperature(start=self.tpt[0], stop=self.tpt[1], num=self.mc_steps):
            self.update(t)
            self.spin_log = np.vstack((self.spin_log, self.spin))
            self.energy_log.append(Potts2D.sys_energy(self.spin, self.cps, self.nbs))

        return self.spin_log, self.energy_log


def graph_coloring(spin_num, spin_value, graph_mode, temperature, MC_steps):
    neighbor_index = generate_neighbor(size=(int(np.sqrt(spin_num)), int(np.sqrt(spin_num))), graph_mode=graph_mode)
    coupling_list = generate_coupling(spin=None, neighbor=neighbor_index, mode="Potts")

    potts = Potts2D(spin_num=spin_num,
                    spin_value=spin_value,
                    couplings=coupling_list,
                    neighbors=neighbor_index,
                    temperature=temperature,
                    MC_steps=MC_steps,
                    Ising_mode=False
                    )
    print("Initial configuration\n", potts.spin, "\n")
    color_map = ListedColormap(['red', 'yellow', 'blue', 'green'])

    spin_log, energy_log = potts.run()
    print("Solution\n", spin_log[-1])
    print("System:", energy_log[-1])
    print()
    print(check(vertexes=spin_log[-1], neighbor_list=neighbor_index))

    row = np.sqrt(spin_num).astype(np.int32)
    col = np.sqrt(spin_num).astype(np.int32)
    plt.figure("color map", figsize=(8, 6))
    # color_map = ListedColormap(['red', 'green', 'blue', 'yellow', 'orange'])
    # color_map = ListedColormap(['red', 'green', 'blue', 'yellow', 'orange', "purple"])
    # plt.ion()
    # for n, spin in enumerate(spins):
    #     plt.cla()
    #     plt.title(f"Iteration: {n}")
    #     plt.imshow(spin.reshape(shape, shape), cmap=color_map)
    #     ticks = np.arange(int(np.sqrt(args.vertexes)))
    #     plt.xticks(ticks, ticks)
    #     plt.yticks(ticks, ticks)
    #     if n < 3 * len(spin):
    #         plt.pause(0.01)
    #     elif n < 5 * len(spin):
    #         plt.pause(0.001)
    #     else:
    #         plt.pause(0.0001)
    # plt.ioff()
    plt.imshow(spin_log[-1].reshape(row, col), cmap=color_map)
    plt.xticks(np.arange(0, row, 2), np.arange(0, row, 2))
    plt.yticks(np.arange(0, col, 2), np.arange(0, col, 2))
    plt.figure("System Energy", figsize=(8, 6))
    plt.plot(np.arange(len(energy_log)), energy_log)
    plt.xlabel("Iteration", fontsize=12)
    plt.ylabel("System Energy", fontsize=12)
    plt.show()


def max_cut(path, graph_mode, temperature, MC_steps):
    row, col = (21, 36), (2, 33)
    shape = (row[1] - row[0] + 1, col[1] - col[0] + 1)
    spin = load_custom(path=path, rows=row, cols=col)

    dim = np.size(spin)
    neighbor_index = generate_neighbor(size=shape, graph_mode=graph_mode)
    coupling_list = generate_coupling(spin, neighbor_index, mode="Ising")

    potts = Potts2D(spin_num=dim,
                    spin_value=np.array([-1, 1]),
                    couplings=coupling_list,
                    neighbors=neighbor_index,
                    temperature=temperature,
                    MC_steps=MC_steps,
                    Ising_mode=True
                    )
    print("Initial configuration\n", potts.spin, "\n")

    spin_log, energy_log = potts.run()
    print("Solution\n", spin_log[-1])
    print("System:", energy_log[-1])
    print()

    plot_spin_config = np.vstack((spin_log[::50, :], spin_log[-1]))
    fig_num = len(plot_spin_config)

    fig = plt.figure("spin fig", figsize=(6, 6))
    plt.tight_layout()
    for n in range(fig_num):
        ax = fig.add_subplot(5, int(fig_num // 5) + 1, n + 1)
        ax.imshow(plot_spin_config[n].reshape(shape), cmap="binary")
        ax.set_xticks([], [])
        ax.set_yticks([], [])

    plt.figure("System Energy", figsize=(8, 6))
    plt.plot(np.arange(len(energy_log)), energy_log)
    plt.xlabel("Iteration", fontsize=12)
    plt.ylabel("System Energy", fontsize=12)
    plt.show()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parameter of Graph Color Map')
    parser.add_argument('--MC_STEP', default=500, help='total iteration')
    parser.add_argument('--spin_num', default=256, help='number of spin')
    parser.add_argument('--temperature', default=(2, 0.125), help='Initial temperature')
    parser.add_argument('--spin_value', default=np.array([0, 1, 2, 3]), help='value of each spin')
    parser.add_argument('--path', default="D:/Lyzenro/Desktop/Sigma.xlsx", help='spin configuration file path')

    args = parser.parse_args()

    # graph_coloring(spin_num=args.spin_num,
    #                spin_value=args.spin_value,
    #                graph_mode="k",
    #                temperature=args.temperature,
    #                MC_steps=args.MC_STEP
    #                )

    max_cut(path=args.path,
            graph_mode="k",
            temperature=args.temperature,
            MC_steps=args.MC_STEP
            )
