import numpy as np
from openpyxl import load_workbook
from matplotlib import pyplot as plt


def load_max_cut(src_path, dst_path, matrix_size):
	coupling = np.zeros(shape=matrix_size, dtype=np.int8)
	with open(src_path, mode="r") as f:
		data = f.read().split("\n")
		print(data)
		for item in data:
			info = item.split(" ")
			i, j, weight = int(info[0]) - 1, int(info[1]) - 1, int(info[2])
			coupling[i, j] = weight
			coupling[j, i] = weight
	np.save(dst_path, coupling)
	print(coupling)
	print(np.array_equal(coupling, coupling.T))


def load_custom(src_path, dst_path, rows: (int, int), cols: (int, int)):
	spins = np.array([])
	start_row, end_row = rows
	start_col, end_col = cols
	sheet = load_workbook(src_path).active

	for i in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
		for j in i:
			spins = np.append(spins, j.value)

	spin_nums = len(spins)
	coupling = np.zeros(shape=(spin_nums, spin_nums), dtype=np.int8)

	for i in range(spin_nums):
		for j in range(i + 1, spin_nums):
			coupling[i, j] = 1 if np.logical_xor(spins[i], spins[j]) else -1
			coupling[j, i] = coupling[i, j]

	np.save(dst_path, coupling)


# print(coupling[:36, :36])
# plt.imshow(spins.reshape(end_row - start_row + 1, end_col - start_col + 1), cmap="binary")
# plt.show()


def load_beasley(src_path, dst_path, matrix_size):
	coupling = np.zeros(shape=matrix_size, dtype=np.int8)
	with open(src_path, mode="r") as f:
		data = f.read().strip().split("\n")
		data.pop(0)
		print(data)

		for item in data:
			info = item.split(" ")
			i, j, weight = int(info[0]) - 1, int(info[1]) - 1, int(info[2])
			coupling[i, j] = weight
			coupling[j, i] = weight
	np.save(dst_path, coupling)
	print(coupling)
	print(np.array_equal(coupling, coupling.T))


if __name__ == '__main__':
	np.set_printoptions(threshold=100000)
	# load_max_cut(src_path="../MaxCutDataset/G61(7000 17148).txt",
	#              dst_path="./CouplingData/G61.npy",
	#              matrix_size=(7000, 7000)
	#              )

	load_custom(src_path="C:/Users/ROBUSTIC/Desktop/Sigma.xlsx",
				dst_path="./CouplingMatrix/Custom_IEEE.npy",
				rows=(23, 38),
				cols=(2, 33)
				)

	# for k in range(1, 11):
	#     load_beasley(src_path=f"../MaxCutDataset/beasley/bqp50-{k}.sparse",
	#                  dst_path=f"./CouplingData/bqp50-{k}.npy",
	#                  matrix_size=(50, 50)
	#                  )
