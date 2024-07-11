import numpy as np
import pandas as pd
from openpyxl import load_workbook

path = r"D:\Lyzenro\Desktop\test.xlsx"
# sheet = load_workbook(path).active
#
# for i in sheet.iter_rows(min_row=23, max_row=38, min_col=2, max_col=33):
#     for j in i:
#         print(type(j.value), j.value)
#         print()
dataframe = pd.read_excel(path)
print(dataframe)
array = dataframe.values

print(array)


